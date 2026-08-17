import os, re, time, json, hashlib, csv, threading
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from fastapi import FastAPI
from pydantic import BaseModel
import uvicorn
from fastembed import TextEmbedding, SparseTextEmbedding
from fastembed.rerank.cross_encoder import TextCrossEncoder
from qdrant_client import QdrantClient
from qdrant_client.models import (
    VectorParams, Distance, SparseVectorParams, SparseIndexParams,
    SparseVector, PointStruct,
)
from neo4j import GraphDatabase
import tiktoken
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# multi-format extractors
import fitz                            # pymupdf — PDF
from docx import Document as DocxDoc   # python-docx — DOCX
import openpyxl                        # XLSX
from lxml import etree                 # SVG
try:
    from pptx import Presentation as PptxPrs  # PPTX (optional)
    _PPTX = True
except ImportError:
    _PPTX = False

VAULT_PATH    = os.environ["VAULT_PATH"]
QDRANT_URL    = os.environ.get("QDRANT_URL", "http://localhost:6333")
NEO4J_URI     = os.environ.get("NEO4J_URI", "bolt://localhost:7687")
NEO4J_USER    = os.environ.get("NEO4J_USER", "neo4j")
NEO4J_PASS    = os.environ.get("NEO4J_PASSWORD", "")
EMBED_PORT    = int(os.environ.get("EMBED_PORT", "7910"))
COLLECTION    = "obsidian_hybrid"
CHUNK_TOKENS  = 400
CHUNK_OVERLAP = 50
STATE_FILE    = "/app/.tapestry-ingest-state.json"
EXTRACT_WORKERS = max(4, (os.cpu_count() or 4))

SUPPORTED = {".md", ".pdf", ".docx", ".xlsx", ".csv", ".pptx", ".svg"}

dense_model  = TextEmbedding("nomic-ai/nomic-embed-text-v1.5")
sparse_model = SparseTextEmbedding("prithivida/Splade_PP_en_v1")
rerank_model = TextCrossEncoder("Xenova/ms-marco-MiniLM-L-6-v2")
enc          = tiktoken.get_encoding("cl100k_base")
qdrant       = QdrantClient(url=QDRANT_URL, timeout=30)
neo4j_driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASS))

# ── State: SHA256-based change detection ──────────────────────────────────────

def _file_hash(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for block in iter(lambda: f.read(65536), b""):
            h.update(block)
    return h.hexdigest()

def _load_state() -> dict:
    try:
        return json.loads(Path(STATE_FILE).read_text())
    except Exception:
        return {}

def _save_state(state: dict):
    try:
        Path(STATE_FILE).write_text(json.dumps(state))
    except Exception as e:
        print(f"[ingest] state save error: {e}")

# ── Format extractors ─────────────────────────────────────────────────────────

def _extract_md(p: Path) -> str:
    return p.read_text(errors="replace")

def _extract_pdf(p: Path) -> str:
    doc = fitz.open(str(p))
    pages = [page.get_text() for page in doc]
    doc.close()
    return "\n\n".join(pages)

def _extract_docx(p: Path) -> str:
    doc = DocxDoc(str(p))
    parts = [para.text for para in doc.paragraphs if para.text.strip()]
    for table in doc.tables:
        for row in table.rows:
            line = " | ".join(c.text.strip() for c in row.cells if c.text.strip())
            if line:
                parts.append(line)
    return "\n".join(parts)

def _extract_xlsx(p: Path) -> str:
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        parts.append(f"Sheet: {name}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)

def _extract_csv(p: Path) -> str:
    with open(p, newline="", errors="replace") as f:
        return "\n".join(" | ".join(row) for row in csv.reader(f))

def _extract_svg(p: Path) -> str:
    tree = etree.parse(str(p))
    tags = {"text", "title", "desc", "tspan"}
    texts = [
        el.text.strip()
        for el in tree.iter()
        if el.tag.split("}")[-1] in tags and el.text and el.text.strip()
    ]
    return " ".join(texts)

def _extract_pptx(p: Path) -> str:
    if not _PPTX:
        return ""
    prs = PptxPrs(str(p))
    parts = []
    for i, slide in enumerate(prs.slides, 1):
        parts.append(f"Slide {i}")
        for shape in slide.shapes:
            if hasattr(shape, "text") and shape.text.strip():
                parts.append(shape.text)
    return "\n".join(parts)

_EXTRACTORS = {
    ".md":   _extract_md,
    ".pdf":  _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
    ".csv":  _extract_csv,
    ".svg":  _extract_svg,
    ".pptx": _extract_pptx,
}

def extract_text(p: Path) -> str:
    fn = _EXTRACTORS.get(p.suffix.lower())
    return fn(p) if fn else ""

# ── Qdrant / Neo4j helpers ────────────────────────────────────────────────────

def ensure_collection():
    names = [c.name for c in qdrant.get_collections().collections]
    if COLLECTION not in names:
        qdrant.create_collection(
            COLLECTION,
            vectors_config={"dense": VectorParams(size=768, distance=Distance.COSINE)},
            sparse_vectors_config={
                "sparse": SparseVectorParams(index=SparseIndexParams(on_disk=False))
            },
        )

def _embed_dense(text: str) -> list[float]:
    return next(dense_model.embed([text])).tolist()

def _embed_sparse(text: str) -> SparseVector:
    sv = next(sparse_model.embed([text]))
    return SparseVector(indices=sv.indices.tolist(), values=sv.values.tolist())

def chunk_text(text: str) -> list[str]:
    toks = enc.encode(text)
    step = CHUNK_TOKENS - CHUNK_OVERLAP
    chunks = []
    for i in range(0, len(toks), step):
        chunks.append(enc.decode(toks[i : i + CHUNK_TOKENS]))
    return chunks or [text[:2000]]

def wikilinks(text: str) -> list[str]:
    return re.findall(r"\[\[([^\]|#]+)", text)

def stable_id(path: str, chunk: int) -> int:
    return abs(hash(f"{path}:{chunk}")) % (2**63)

def index_file(p: Path):
    rel  = str(p.relative_to(VAULT_PATH))
    text = extract_text(p)
    if not text.strip():
        return
    chunks = chunk_text(text)
    points = []
    for i, chunk in enumerate(chunks):
        dv = _embed_dense(chunk)
        sv = _embed_sparse(chunk)
        points.append(PointStruct(
            id=stable_id(rel, i),
            vector={"dense": dv, "sparse": sv},
            payload={"path": rel, "chunk": i, "text": chunk},
        ))
    qdrant.upsert(collection_name=COLLECTION, points=points, wait=False)
    if p.suffix.lower() == ".md":
        links = wikilinks(text)
        with neo4j_driver.session(database="fde-rag") as s:
            s.run("MERGE (n:Note {path:$p})", p=rel)
            for lnk in links:
                s.run(
                    "MERGE (n:Note {path:$p}) "
                    "MERGE (m:Note {path:$m}) "
                    "MERGE (n)-[:LINKS_TO]->(m)",
                    p=rel, m=lnk + ".md",
                )

# ── Watchdog ──────────────────────────────────────────────────────────────────

class VaultHandler(FileSystemEventHandler):
    def _handle(self, src: str):
        p = Path(src)
        if not p.is_file() or p.suffix.lower() not in SUPPORTED:
            return
        try:
            index_file(p)
        except Exception as e:
            print(f"[ingest] error {src}: {e}")

    def on_modified(self, ev):
        if not ev.is_directory:
            self._handle(ev.src_path)

    on_created = on_modified

# ── Initial vault scan (parallel extraction, serial embedding) ────────────────

def run_watcher():
    ensure_collection()
    state = _load_state()

    all_files = [
        p for p in Path(VAULT_PATH).rglob("*")
        if p.is_file() and p.suffix.lower() in SUPPORTED
    ]

    # parallel extraction — I/O + decompression bound
    def _extract(p: Path):
        h = _file_hash(p)
        if state.get(str(p)) == h:
            return p, None, h          # unchanged — skip
        return p, extract_text(p), h

    changed = []
    with ThreadPoolExecutor(max_workers=EXTRACT_WORKERS) as pool:
        futures = {pool.submit(_extract, p): p for p in all_files}
        for fut in as_completed(futures):
            try:
                p, text, h = fut.result()
                if text is not None:
                    changed.append((p, text, h))
            except Exception as e:
                print(f"[ingest] extract error {futures[fut]}: {e}")

    print(f"[ingest] {len(changed)}/{len(all_files)} files need indexing")

    # serial embedding (ONNX is already multi-threaded internally)
    for p, text, h in changed:
        try:
            chunks = chunk_text(text)
            points = []
            for i, chunk in enumerate(chunks):
                dv = _embed_dense(chunk)
                sv = _embed_sparse(chunk)
                points.append(PointStruct(
                    id=stable_id(str(p.relative_to(VAULT_PATH)), i),
                    vector={"dense": dv, "sparse": sv},
                    payload={
                        "path": str(p.relative_to(VAULT_PATH)),
                        "chunk": i,
                        "text": chunk,
                    },
                ))
            qdrant.upsert(collection_name=COLLECTION, points=points, wait=False)
            if p.suffix.lower() == ".md":
                links = wikilinks(text)
                rel = str(p.relative_to(VAULT_PATH))
                with neo4j_driver.session(database="fde-rag") as s:
                    s.run("MERGE (n:Note {path:$p})", p=rel)
                    for lnk in links:
                        s.run(
                            "MERGE (n:Note {path:$p}) "
                            "MERGE (m:Note {path:$m}) "
                            "MERGE (n)-[:LINKS_TO]->(m)",
                            p=rel, m=lnk + ".md",
                        )
            state[str(p)] = h
        except Exception as e:
            print(f"[ingest] index error {p}: {e}")

    _save_state(state)
    print(f"[ingest] initial scan complete — {len(changed)} files indexed")

    obs = Observer()
    obs.schedule(VaultHandler(), VAULT_PATH, recursive=True)
    obs.start()
    print("[ingest] watcher running")
    try:
        while True:
            time.sleep(10)
    except KeyboardInterrupt:
        obs.stop()
    obs.join()

# ── FastAPI embed API ─────────────────────────────────────────────────────────

app = FastAPI()

class EmbedReq(BaseModel):
    text: str

class RerankReq(BaseModel):
    query: str
    passages: list[str]

@app.get("/health")
def health():
    return {"ok": True}

@app.post("/embed")
def embed(req: EmbedReq):
    return {"vector": _embed_dense(req.text)}

@app.post("/embed-sparse")
def embed_sparse(req: EmbedReq):
    sv = _embed_sparse(req.text)
    return {"indices": sv.indices, "values": sv.values}

@app.post("/rerank")
def rerank(req: RerankReq):
    scores = list(rerank_model.rerank(req.query, req.passages))
    return {"scores": [float(s) for s in scores]}

if __name__ == "__main__":
    t = threading.Thread(target=run_watcher, daemon=True)
    t.start()
    uvicorn.run(app, host="0.0.0.0", port=EMBED_PORT)
